import os
import sqlite3
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


DATABASE = "db/nifty100.db"


class PeerReport:

    def __init__(self):

        self.conn = sqlite3.connect(DATABASE)

        self.green = PatternFill(
            fill_type="solid",
            start_color="92D050"
        )

        self.yellow = PatternFill(
            fill_type="solid",
            start_color="FFD966"
        )

        self.red = PatternFill(
            fill_type="solid",
            start_color="FF9999"
        )

        self.gold = PatternFill(
            fill_type="solid",
            start_color="FFC000"
        )

        self.bold = Font(bold=True)

    ###########################################################

    def load_data(self):

        query = """

        SELECT

            pg.peer_group_name,

            c.id AS company_id,
            c.company_name,
            c.roce_percentage,

            fr.year,

            fr.return_on_equity_pct,
            fr.net_profit_margin_pct,
            fr.operating_profit_margin_pct,

            fr.debt_to_equity,
            fr.interest_coverage,
            fr.asset_turnover,

            fr.free_cash_flow_cr,
            fr.capex_cr,

            fr.cash_from_operations_cr,
            fr.total_debt_cr,

            fr.earnings_per_share,
            fr.book_value_per_share,
            fr.dividend_payout_ratio_pct,

            fr.revenue_cagr_3yr,
            fr.revenue_cagr_5yr,
            fr.pat_cagr_5yr,
            fr.eps_cagr_5yr,

            fr.composite_quality_score,

            pp.metric,
            pp.percentile_rank

        FROM peer_groups pg

        LEFT JOIN companies c
            ON pg.company_id = c.id

        LEFT JOIN financial_ratios fr
            ON pg.company_id = fr.company_id

        LEFT JOIN peer_percentiles pp
            ON
                fr.company_id = pp.company_id
                AND fr.year = pp.year

        ORDER BY
            pg.peer_group_name,
            c.company_name,
            fr.year

        """

        df = pd.read_sql(query, self.conn)

        return df

    ###########################################################

    def create_workbook(self, df):

        os.makedirs(
            "output",
            exist_ok=True
        )

        wb = Workbook()

        wb.remove(wb.active)

        peer_groups = sorted(

            df["peer_group_name"]
            .dropna()
            .unique()

        )

        print(
            f"Peer Groups Found : {len(peer_groups)}"
        )

        for group in peer_groups:

            ws = wb.create_sheet(

                title=group[:31]

            )

            group_df = df[
                df["peer_group_name"] == group
            ].copy()

            group_df = group_df.sort_values(

                by="composite_quality_score",

                ascending=False

            )

            headers = [

                "company_id",
                "company_name",
                "year",

                "return_on_equity_pct",
                "roce_percentage",
                "net_profit_margin_pct",
                "operating_profit_margin_pct",

                "debt_to_equity",
                "interest_coverage",
                "asset_turnover",

                "free_cash_flow_cr",
                "capex_cr",
                "cash_from_operations_cr",
                "total_debt_cr",

                "earnings_per_share",
                "book_value_per_share",
                "dividend_payout_ratio_pct",

                "revenue_cagr_3yr",
                "revenue_cagr_5yr",
                "pat_cagr_5yr",
                "eps_cagr_5yr",

                "composite_quality_score",

                "metric",
                "percentile_rank"

            ]

            ws.append(headers)

            for cell in ws[1]:

                cell.font = self.bold

                cell.alignment = Alignment(
                    horizontal="center"
                )

            for row in group_df[headers].itertuples(index=False):

                ws.append(list(row))
            ####################################################
            # Colour percentile rank
            ####################################################

            if "percentile_rank" in headers:

                percentile_col = (
                    headers.index("percentile_rank") + 1
                )

                for row in range(2, ws.max_row + 1):

                    value = ws.cell(
                        row=row,
                        column=percentile_col
                    ).value

                    if value is None:
                        continue

                    try:

                        value = float(value)

                    except:

                        continue

                    if value >= 75:

                        ws.cell(
                            row=row,
                            column=percentile_col
                        ).fill = self.green

                    elif value <= 25:

                        ws.cell(
                            row=row,
                            column=percentile_col
                        ).fill = self.red

                    else:

                        ws.cell(
                            row=row,
                            column=percentile_col
                        ).fill = self.yellow

            ####################################################
            # Benchmark Row
            ####################################################

            if ws.max_row >= 2:

                for cell in ws[2]:

                    cell.fill = self.gold

                    cell.font = self.bold

            ####################################################
            # Median Row
            ####################################################

            numeric_columns = group_df.select_dtypes(
                include="number"
            ).columns

            median_row = []

            for col in headers:

                if col == "company_name":

                    median_row.append("Peer Median")

                elif col in numeric_columns:

                    median_row.append(
                        round(group_df[col].median(), 2)
                    )

                else:

                    median_row.append("")

            ws.append(median_row)

            last_row = ws.max_row

            for cell in ws[last_row]:

                cell.font = self.bold

            ####################################################
            # Auto Fit Columns
            ####################################################

            for column_cells in ws.columns:

                length = max(

                    len(str(cell.value))
                    if cell.value is not None
                    else 0

                    for cell in column_cells

                )

                ws.column_dimensions[
                    column_cells[0].column_letter
                ].width = min(length + 3, 35)

        ####################################################
        # Save Workbook
        ####################################################

        wb.save(
            "output/peer_comparison.xlsx"
        )

        print(
            "\npeer_comparison.xlsx created successfully."
        )

    ###########################################################

    def close(self):

        self.conn.close()