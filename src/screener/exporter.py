import os

from openpyxl import Workbook
from openpyxl.styles import PatternFill

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


class ScreenerExporter:

    def __init__(self):
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

    def apply_colours(self, ws, columns, preset_name):

        rules = {
            "quality_compounder": {
                "return_on_equity_pct": (15, ">="),
                "debt_to_equity": (1, "<="),
                "free_cash_flow_cr": (0, ">="),
                "revenue_cagr_5yr": (10, ">="),
            },
            "value_pick": {
                "pe_ratio": (20, "<="),
                "pb_ratio": (3, "<="),
                "dividend_yield_pct": (1, ">="),
            },
            "growth_accelerator": {
                "revenue_cagr_5yr": (15, ">="),
                "pat_cagr_5yr": (20, ">="),
            },
            "dividend_champion": {
                "dividend_yield_pct": (4, ">="),
                "return_on_equity_pct": (18, ">="),
            },
            "debt_free_blue_chip": {"debt_to_equity": (0, "<=")},
            "turnaround_watch": {
                "revenue_cagr_5yr": (20, ">="),
                "free_cash_flow_cr": (100, ">="),
            },
        }

        if preset_name not in rules:
            return

        preset = rules[preset_name]

        for row in ws.iter_rows(min_row=2):

            for cell in row:

                column = columns[cell.column - 1]

                if column not in preset:
                    continue

                value, operator = preset[column]

                if not isinstance(cell.value, (int, float)):
                    continue

                passed = False

                if operator == ">=":
                    passed = cell.value >= value

                elif operator == "<=":
                    passed = cell.value <= value

                if passed:
                    cell.fill = GREEN
                else:
                    cell.fill = RED

    def export(self, results):

        for preset_name, df in results.items():

            ws = self.workbook.create_sheet(title=preset_name[:31])

            # Select only required columns
            columns = [
                "company_name",
                "broad_sector",
                "sub_sector",
                "return_on_equity_pct",
                "roce_percentage",
                "net_profit_margin_pct",
                "free_cash_flow_cr",
                "revenue_cagr_5yr",
                "pat_cagr_5yr",
                "eps_cagr_5yr",
                "debt_to_equity",
                "interest_coverage",
                "asset_turnover",
                "sales",
                "net_profit",
                "operating_profit",
                "market_cap_crore",
                "pe_ratio",
                "pb_ratio",
                "composite_quality_score",
            ]

            df = df[columns]

            # Write headers
            ws.append(columns)
            # Write data
            for row in df.itertuples(index=False):
                ws.append(list(row))

            self.apply_colours(ws, columns, preset_name)
        os.makedirs("output", exist_ok=True)

        self.workbook.save("output/screener_output.xlsx")

        print("Excel exported successfully.")
